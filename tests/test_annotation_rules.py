"""Tests for dmpbridge.evaluation.annotation_rules (Path B: new-annotation eval).

``_RULES`` is a transcription of ``data/input/Rules.xlsx``, so the tests here
check two separate things:

* that the transcription matches the spreadsheet, row for row — the one test
  that will catch the sheet and the code drifting apart;
* that each row's action is carried out literally — an empty question is filled
  from section.title, then section.description, then the document title, and a
  question that already has text is left alone.

convert_tag_to_final() and load_method_new() are tested against temporary
directories via monkeypatch so they don't touch real project data.
"""
import json
import re

import openpyxl
import pytest

import dmpbridge.evaluation.annotation_rules as ar

FIELDS = ("title", "section.title", "section.description", "question.text")


# ── The table matches the spreadsheet ────────────────────────────────────────

def test_rules_table_matches_the_spreadsheet():
    """The authoritative check: _RULES must transcribe Rules.xlsx exactly.

    If this fails, either the sheet changed and _RULES needs re-transcribing, or
    someone edited _RULES directly — which would silently diverge from the spec.
    """
    sheet = openpyxl.load_workbook("data/input/Rules.xlsx", data_only=True).worksheets[0]

    # The sheet's column order has changed before, so read it from the header
    # rather than assuming it — a reordered sheet must not silently remap rows.
    header = [str(c).strip() if c else "" for c in next(sheet.iter_rows(values_only=True))]
    assert header[1:5] == list(ar.RULE_FIELDS), (
        f"Rules.xlsx column order is {header[1:5]}, but RULE_FIELDS is "
        f"{list(ar.RULE_FIELDS)} — re-transcribe _RULES against the new order")

    seen = 0
    for row in sheet.iter_rows(min_row=2, max_row=17, values_only=True):
        n, action = row[0], row[5]
        key = tuple(row[1:5])
        m = re.search(r'Copy "?([\w.]+)"? into "?([\w.]+)"?', action or "")
        expected = (m.group(1), m.group(2)) if m else None
        assert ar._RULES[key] == expected, f"row {n}: sheet says {expected}, code has {ar._RULES[key]}"
        seen += 1
    assert seen == 16, "expected 16 rule rows in the spreadsheet"


def test_every_emptiness_combination_is_covered():
    """All 2^4 patterns must be present, so no input can fall through."""
    assert len(ar._RULES) == 16
    for i in range(16):
        key = tuple("EN"[(i >> (3 - b)) & 1] for b in range(4))
        assert key in ar._RULES


# ── Helpers ──────────────────────────────────────────────────────────────────

def _doc(title="", sec_title="", sec_desc="", q_text="", answer="content"):
    return {"narrative": {"template": {"title": title, "section": [{
        "title": sec_title,
        "description": sec_desc,
        "question": [{"text": q_text, "answer": {"json": {"answer": answer}}}],
    }]}}}


def _fields(out):
    t = out["narrative"]["template"]
    s = t["section"][0]
    return t["title"], s["title"], s["question"][0]["text"]


# ── Filling an empty question ────────────────────────────────────────────────
#
# Every odd row has question.text empty and fills it from the first available
# source, in order: section.title > section.description > document title.

def test_row_1_leaves_the_question_empty_when_nothing_is_available():
    """E/E/E/E — no source to copy from."""
    _, sec_title, q_text = _fields(ar.apply_new_annotation_rules(_doc()))
    assert (sec_title, q_text) == ("", "")


def test_rows_3_and_11_fall_through_to_the_description():
    """E/E/N/E and N/E/N/E — no section title, so the description fills the question."""
    for title in ("", "Doc Title"):
        _, sec_title, q_text = _fields(ar.apply_new_annotation_rules(
            _doc(title=title, sec_title="", sec_desc="Funder guidance", q_text="")))
        assert q_text == "Funder guidance"
        assert sec_title == ""


def test_rows_5_7_13_15_prefer_the_section_title():
    """All four patterns where section.title is present and question.text is empty."""
    for title in ("", "Doc Title"):
        for sec_desc in ("", "Funder guidance"):
            _, sec_title, q_text = _fields(ar.apply_new_annotation_rules(
                _doc(title=title, sec_title="Sec 1", sec_desc=sec_desc, q_text="")))
            assert q_text == "Sec 1", f"{title=} {sec_desc=}"
            assert sec_title == "Sec 1"


def test_row_9_falls_through_to_the_document_title():
    """N/E/E/E — the only row that reads the document title."""
    title, sec_title, q_text = _fields(ar.apply_new_annotation_rules(_doc(title="Doc Title")))
    assert q_text == "Doc Title"
    assert sec_title == ""          # no row writes to section.title
    assert title == "Doc Title"     # the title itself is never modified


def test_section_title_outranks_description_and_document_title():
    """Row 15 (N/N/N/E) — all three sources present, section title wins."""
    _, _, q_text = _fields(ar.apply_new_annotation_rules(
        _doc(title="Doc Title", sec_title="Sec 1", sec_desc="Guidance", q_text="")))
    assert q_text == "Sec 1"


def test_description_outranks_the_document_title():
    """Row 11 (N/E/N/E) — no section title, so the description beats the title."""
    _, _, q_text = _fields(ar.apply_new_annotation_rules(
        _doc(title="Doc Title", sec_title="", sec_desc="Guidance", q_text="")))
    assert q_text == "Guidance"


# ── Leaving a populated question alone ───────────────────────────────────────

def test_existing_question_text_is_never_touched():
    """Every even row: question.text already has text, so nothing changes.

    This covers all eight combinations of the other three fields.
    """
    for title in ("", "Doc Title"):
        for sec_title in ("", "Sec 1"):
            for sec_desc in ("", "Guidance"):
                _, got_sec, got_q = _fields(ar.apply_new_annotation_rules(
                    _doc(title=title, sec_title=sec_title, sec_desc=sec_desc,
                         q_text="Original question")))
                assert got_q == "Original question", f"{title=} {sec_title=} {sec_desc=}"
                assert got_sec == sec_title


# ── Invariants across the whole table ────────────────────────────────────────

def test_section_title_is_never_written():
    """No row in this revision targets section.title."""
    assert all(a is None or a[1] != "section.title" for a in ar._RULES.values())


def test_document_title_is_never_modified():
    for title in ("", "Doc Title"):
        for sec_title in ("", "Sec 1"):
            for q_text in ("", "Q"):
                for sec_desc in ("", "Desc"):
                    out = ar.apply_new_annotation_rules(
                        _doc(title=title, sec_title=sec_title,
                             q_text=q_text, sec_desc=sec_desc))
                    assert out["narrative"]["template"]["title"] == title


def test_input_is_not_mutated():
    data = _doc(title="Doc", sec_title="Sec 1", q_text="Original")
    ar.apply_new_annotation_rules(data)
    assert data["narrative"]["template"]["section"][0]["question"][0]["text"] == "Original"


def test_document_title_fills_every_matching_question():
    """The sheet states no once-only restriction, so every N/E/E/E question is filled."""
    data = {"narrative": {"template": {"title": "Doc Title", "section": [
        {"title": "", "description": "",
         "question": [{"text": "", "answer": {"json": {"answer": "a"}}}]},
        {"title": "", "description": "",
         "question": [{"text": "", "answer": {"json": {"answer": "b"}}}]},
    ]}}}
    sections = ar.apply_new_annotation_rules(data)["narrative"]["template"]["section"]
    assert sections[0]["question"][0]["text"] == "Doc Title"
    assert sections[1]["question"][0]["text"] == "Doc Title"


# ── Behaviour on the real reference data ─────────────────────────────────────

def test_reproduces_the_new_annotation_for_every_sample():
    """The rules reproduce the new-version reference annotation on all 10 documents.

    This is the strongest available check that the transcription and the sheet's
    column order are both right — a mis-ordered column drops this sharply.
    """
    from dmpbridge.evaluation.evaluate import resolve_old_gt_path

    def pairs(doc):
        t = doc["narrative"]["template"]
        return [((s.get("title") or "").strip(), (q.get("text") or "").strip())
                for s in t.get("section", []) for q in s.get("question", [])]

    matching = []
    for n in range(1, 11):
        old = json.loads(resolve_old_gt_path(n).read_text(encoding="utf-8"))
        new = json.loads(ar.resolve_new_gt_path(n).read_text(encoding="utf-8"))
        if pairs(ar.apply_new_annotation_rules(old)) == pairs(new):
            matching.append(n)
    assert matching == list(range(1, 11)), f"no longer reproduces every sample: {matching}"


# ── resolve_new_gt_path ──────────────────────────────────────────────────────

def test_resolve_new_gt_path_finds_both_naming_patterns():
    # samples 1,2,3,4,7 use sampleN_dmp_new.json; samples 5,6,8,9,10 use dmp_sampleN_new.json
    for n in [1, 5]:
        path = ar.resolve_new_gt_path(n)
        assert path.exists()
        assert str(n) in path.stem


def test_resolve_new_gt_path_raises_for_unknown_sample():
    with pytest.raises(FileNotFoundError):
        ar.resolve_new_gt_path(9999)


# ── convert_tag_to_final / load_method_new (isolated via monkeypatch) ───────

def _write_structured(path, title, section_title, question_text, answer):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps({"narrative": {"template": {
        "title": title,
        "section": [{
            "title": section_title,
            "question": [{"text": question_text, "answer": {"json": {"answer": answer}}}],
        }],
    }}}), encoding="utf-8")


def test_convert_tag_to_final_writes_rule_converted_output(tmp_path, monkeypatch):
    """Stage 3 -> stage 4, with identical filenames in each stage directory."""
    structured_dir = tmp_path / "3_structured"
    final_dir      = tmp_path / "4_final"
    monkeypatch.setattr(ar._paths, "STRUCTURED_DIR", structured_dir)
    monkeypatch.setattr(ar, "FINAL_DIR", final_dir)

    # N/E/E/E -> row 9: the document title fills the question.
    _write_structured(structured_dir / "my-tag" / "sample1.json",
                       title="Doc", section_title="", question_text="", answer="content")

    n = ar.convert_tag_to_final("my-tag")

    assert n == 1
    out = json.loads((final_dir / "my-tag" / "sample1.json").read_text(encoding="utf-8"))
    assert out["narrative"]["template"]["section"][0]["question"][0]["text"] == "Doc"


def test_load_method_scores_final_json_against_resolved_gold(tmp_path, monkeypatch):
    final_dir = tmp_path / "4_final"
    monkeypatch.setattr(ar, "FINAL_DIR", final_dir)

    _write_structured(final_dir / "my-tag" / "sample1.json",
                       title="Doc", section_title="Sec 1",
                       question_text="Sec 1", answer="The answer text.")

    # load_method_new drives from PATH_B.annotation_dir (a real, listable
    # directory — this is what makes an EvaluationPath declarable as plain
    # data, e.g. from YAML), not from a per-sample resolver function. So the
    # gold file is written directly, and PATH_B is repointed at tmp_path.
    gold_path = tmp_path / "gold1.json"
    gold_path.write_text(json.dumps({"narrative": {"template": {
        "title": "Doc",
        "section": [{
            "title": "Sec 1",
            "question": [{"text": "Sec 1", "answer": {"json": {"answer": "The answer text."}}}],
        }],
    }}}), encoding="utf-8")

    # Both fields need overriding: PATH_B.predicted_dir was already snapshotted
    # to the real FINAL_DIR value when evaluate.py was first imported, so
    # monkeypatching the *name* FINAL_DIR above does not retroactively change it.
    monkeypatch.setattr(
        ar, "PATH_B",
        ar.replace(ar.PATH_B, annotation_dir=tmp_path, predicted_dir=final_dir),
    )

    df, conf, errors = ar.load_method_new("my-tag")

    assert df is not None
    assert int(df["correct"].sum()) == int(df["total"].sum())  # perfect match
    assert conf["answer.text"]["answer.text"] == 1


def test_load_method_returns_none_for_missing_tag(tmp_path, monkeypatch):
    monkeypatch.setattr(ar, "FINAL_DIR", tmp_path / "4_final")
    df, conf, errors = ar.load_method_new("does-not-exist")
    assert df is None
    assert conf is None
    assert errors is None
